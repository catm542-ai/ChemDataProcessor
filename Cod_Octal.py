import pandas as pd
import re
from openpyxl import load_workbook
import numpy as np
from PIL import Image

# Step 1: Transform the text file into an Excel file (fill empty cells with 0)
# Set the input text file name here:
input_text_file = 'octal_def_lowlevel.txt'
data = []
compound_name = None

# Regular expressions to match compound names and data rows
compound_pattern = re.compile(r"Compound \d+:  (.+)")
data_pattern = re.compile(r"(\d+)\s+(\d+)\s+(\S+)\s+(\S+)\s+([\d\.]*)\s+([\d\.]*)\s+([\d\.]*)\s*([\d\.]*)\s+(\S*)\s*([\d\.\-]*)\s*([\d\.\-]*)")

# Read the input text file and extract data
with open(input_text_file, 'r') as file:
    for line in file:
        compound_match = compound_pattern.match(line)
        if compound_match:
            compound_name = compound_match.group(1).strip().lower()
            continue

        match = data_pattern.match(line)
        if match:
            row = [compound_name] + list(match.groups())
            data.append(row)

# Define column names for the DataFrame
columns = ["Compound", "#", "ID", "Type", "Std. Conc", "RT", "Area", "IS Area", "Response", "Detection Flags", "pg on column", "%Dev"]
df_transformed = pd.DataFrame(data, columns=columns)

# Convert numeric columns and fill empty cells with 0
for col in ["#", "Std. Conc", "RT", "Area", "IS Area", "Response", "pg on column", "%Dev"]:
    df_transformed[col] = pd.to_numeric(df_transformed[col], errors='coerce').fillna(0)
df_transformed = df_transformed.fillna(0)

# Save the transformed data to an Excel file
# Set the output Excel file name for the transformed data:
transformed_excel_path = 'transformed_compounds.xlsx'
df_transformed.to_excel(transformed_excel_path, index=False)
print(f"Transformed file saved to {transformed_excel_path}")

# Step 2: Load the original Excel file and extract compound names
# Set the input Excel file name here:
excel_file_path = 'octal.xlsx'
df_original = pd.read_excel(excel_file_path, header=None)

# Lists to store compound names and their locations in the original file
compounds = []
compound_locations = []
compound_found = False

# Search for the "compound" column in the original Excel file
for index, row in df_original.iterrows():
    for col_index, cell in enumerate(row):
        if isinstance(cell, str) and cell.strip().lower() == 'compound':
            compound_found = True
            break

    if compound_found and pd.notna(row[1]):
        compound_name = row[1]
        if isinstance(compound_name, str) and compound_name.lower().strip() != 'compound':
            compounds.append(compound_name.strip().lower())  # Clean spaces and convert to lowercase
            compound_locations.append((index, 1))

# Step 3: Extract and classify the seventh vial from the transformed data
df_transformed = pd.read_excel(transformed_excel_path)
concentration_data = {}

for compound in compounds:
    # Filter rows by compound name and select the seventh vial
    filtered_rows = df_transformed[(df_transformed['Compound'] == compound) & 
                                   (df_transformed['#'] == 7)]
    if not filtered_rows.empty:
        concentration = filtered_rows.iloc[0]['pg on column']  # Get the seventh vial's value
    else:
        concentration = 0  # Assume 0 if no data is found

    # Classify the value based on the specified ranges
    if concentration < 50:
        classification = 0
    elif 50 <= concentration < 80:
        classification = 1
    elif 80 <= concentration < 210:
        classification = 2
    elif 210 <= concentration < 350:
        classification = 3
    elif 350 <= concentration < 475:
        classification = 4
    elif 475 <= concentration < 700:
        classification = 5
    elif 700 <= concentration < 900:
        classification = 6
    else:  # Concentration >= 900
        classification = 7

    concentration_data[compound] = classification

# Step 4: Write the classified values to the original Excel file
# Load the original workbook and active sheet
workbook = load_workbook(excel_file_path)
sheet = workbook.active  # Assume the first active sheet

# Create a matrix for visualization and write the classified values
matrix_data = []
for compound, (row_index, col_index) in zip(compounds, compound_locations):
    # Get the classification; write 0 if the compound is missing
    classification = concentration_data.get(compound, 0)
    matrix_data.append([classification])  # Only one value per compound

    # Write the classification to the appropriate column
    sheet.cell(row=row_index + 1, column=col_index + 2, value=classification)

# Save the updated Excel file
workbook.save(excel_file_path)
print("Excel file updated with classified seventh vial values.")

# Step 5: Create an image based on the classification matrix
# Convert the matrix list into a numpy array
matrix_array = np.array(matrix_data)

# Define colors for each classification
color_map = {
    0: [0, 0, 0],       # Black
    1: [0, 255, 0],     # Green
    2: [0, 0, 255],     # Blue
    3: [255, 255, 0],   # Yellow
    4: [255, 165, 0],   # Orange
    5: [255, 0, 0],     # Red
    6: [128, 0, 128],   # Purple
    7: [255, 255, 255]  # White
}

# Create the image from the classification data
image_data = np.zeros((matrix_array.shape[0], 1, 3), dtype=np.uint8)
for i in range(matrix_array.shape[0]):
    image_data[i, 0] = color_map[matrix_array[i, 0]]

# Scale the image to make it larger
scale_factor = 50  # Adjust scale factor as needed
large_image = Image.fromarray(image_data, 'RGB').resize(
    (scale_factor, image_data.shape[0] * scale_factor), Image.NEAREST
)

# Save the enlarged image
large_image.save("seventh_vial_classification.png")
print("Enlarged image saved as seventh_vial_classification.png")

# Create and save the rotated and flipped image
corrected_image = large_image.rotate(-90, expand=True).transpose(Image.FLIP_LEFT_RIGHT)
corrected_image.save("seventh_vial_classification_corrected.png")
print("Corrected image saved as seventh_vial_classification_corrected.png")
