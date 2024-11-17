import pandas as pd
import re
from openpyxl import load_workbook
import numpy as np
from PIL import Image

# Step 1: Transform the text file into an Excel file (filling empty cells with 0)
# Specify the input text file containing the raw data
input_text_file = 'quaternary.txt'  # Replace with your input file name
data = []
compound_name = None

# Regular expressions to match compound names and data rows
compound_pattern = re.compile(r"Compound \d+:  (.+)")  # Matches compound headers like "Compound 1: ..."
data_pattern = re.compile(r"(\d+)\s+(\d+)\s+(\S+)\s+(\S+)\s+([\d\.]*)\s+([\d\.]*)\s+([\d\.]*)\s*([\d\.]*)\s+(\S*)\s*([\d\.\-]*)\s*([\d\.\-]*)")  # Matches data rows

# Parse the input text file and extract data
with open(input_text_file, 'r') as file:
    for line in file:
        compound_match = compound_pattern.match(line)
        if compound_match:
            compound_name = compound_match.group(1).strip().lower()
            continue

        match = data_pattern.match(line)
        if match:
            row = [compound_name] + list(match.groups())
            data.append(row)

# Define column names for the DataFrame
columns = ["Compound", "#", "ID", "Type", "Std. Conc", "RT", "Area", "IS Area", "Response", "Detection Flags", "pg on column", "%Dev"]
df_transformed = pd.DataFrame(data, columns=columns)

# Convert numeric columns and fill empty cells with 0
for col in ["#", "Std. Conc", "RT", "Area", "IS Area", "Response", "pg on column", "%Dev"]:
    df_transformed[col] = pd.to_numeric(df_transformed[col], errors='coerce').fillna(0)
df_transformed = df_transformed.fillna(0)

# Save the transformed data to an Excel file
# Specify the output Excel file name
transformed_excel_path = 'transformed_compounds.xlsx'  # Replace with your desired file name
df_transformed.to_excel(transformed_excel_path, index=False)
print(f"Transformed file saved to {transformed_excel_path}")

# Step 2: Load the original Excel file and extract compound names
# Specify the input Excel file containing the original data
excel_file_path = 'quaternary.xlsx'  # Replace with your input file name
df_original = pd.read_excel(excel_file_path, header=None)

# Lists to store compound names and their locations in the original file
compounds = []
compound_locations = []
compound_found = False

# Search for the "Compound" header in the original Excel file
for index, row in df_original.iterrows():
    for col_index, cell in enumerate(row):
        if isinstance(cell, str) and cell.strip().lower() == 'compound':
            compound_found = True
            break

    if compound_found and pd.notna(row[1]):
        compound_name = row[1]
        if isinstance(compound_name, str) and compound_name.lower().strip() != 'compound':
            compounds.append(compound_name.strip().lower())  # Clean spaces and convert to lowercase
            compound_locations.append((index, 1))

# Step 3: Extract and classify concentration values for vials 7-19
df_transformed = pd.read_excel(transformed_excel_path)
concentration_data = {}

for compound in compounds:
    # Filter rows by compound name and vials 7-19
    filtered_rows = df_transformed[(df_transformed['Compound'] == compound) & 
                                   (df_transformed['#'].between(7, 19))]
    # Extract concentration values
    concentrations = filtered_rows['pg on column'].tolist()
    
    # Classify concentrations into predefined ranges
    classified_concentrations = []
    for conc in concentrations:
        if conc < 70:
            classified_concentrations.append(0)
        elif 70 <= conc <= 250:
            classified_concentrations.append(1)
        elif 250 < conc <= 700:
            classified_concentrations.append(2)
        else:
            classified_concentrations.append(3)

    # Ensure exactly 10 values per compound, filling with 0s if fewer
    if len(classified_concentrations) < 10:
        classified_concentrations += [0] * (10 - len(classified_concentrations))
    elif len(classified_concentrations) > 10:
        classified_concentrations = classified_concentrations[:10]  # Limit to 10 values

    concentration_data[compound] = classified_concentrations

# Step 4: Write classified values to the original Excel file
# Open the original workbook
workbook = load_workbook(excel_file_path)
sheet = workbook.active  # Use the first active sheet

# Create a data matrix for visualization and update the Excel file
matrix_data = []
for compound, (row_index, col_index) in zip(compounds, compound_locations):
    # Get classified concentrations or fill with 10 zeros if compound is missing
    classified_concentrations = concentration_data.get(compound, [0] * 10)
    matrix_data.append(classified_concentrations)

    # Write each classified value to the corresponding columns
    for i in range(10):
        sheet.cell(row=row_index + 1, column=col_index + 2 + i, value=classified_concentrations[i])

# Save the updated Excel file
workbook.save(excel_file_path)
print("Excel file updated with classified concentration values.")

# Step 5: Create a heatmap image from the classification matrix
matrix_array = np.array(matrix_data)

# Define colors for each classification
color_map = {
    0: [0, 0, 0],       # Black
    1: [0, 255, 0],     # Green
    2: [255, 255, 0],   # Yellow
    3: [255, 0, 0]      # Red
}

# Generate the heatmap image based on classification data
image_data = np.zeros((matrix_array.shape[0], matrix_array.shape[1], 3), dtype=np.uint8)
for i in range(matrix_array.shape[0]):
    for j in range(matrix_array.shape[1]):
        image_data[i, j] = color_map[matrix_array[i, j]]

# Scale the image for better visibility
scale_factor = 20  # Adjust scale factor as needed
large_image = Image.fromarray(image_data, 'RGB').resize(
    (image_data.shape[1] * scale_factor, image_data.shape[0] * scale_factor), Image.NEAREST
)

# Save the scaled heatmap image
large_image.save("concentration_classification.png")
print("Heatmap image saved as concentration_classification.png")

# Create and save a rotated and flipped version of the heatmap
corrected_image = large_image.rotate(-90, expand=True).transpose(Image.FLIP_LEFT_RIGHT)
corrected_image.save("concentration_classification_corrected.png")
print("Corrected heatmap image saved as concentration_classification_corrected.png")
