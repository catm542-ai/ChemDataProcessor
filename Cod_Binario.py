import pandas as pd
import re
import difflib
from openpyxl import load_workbook

# File paths (update these to match your file locations)
excel_file_path = "code.xlsx"  # Input Excel file containing compound names
text_file_path = "32x5_CODE0.txt"  # Input text file with vial data
output_excel_file = "code_updated_with_vials_and_binary.xlsx"  # Output Excel file with binary vial data
output_text_file = "concatenated_binary_codes.txt"  # Output text file with concatenated binary codes

# Load the Excel file to extract compound names
df_excel = pd.read_excel(excel_file_path, header=None)
compound_column_index = 1  # Assuming compound names are in the second column

# Read the input text file containing vial data
with open(text_file_path, "r") as file:
    text_data = file.read()

# Load the workbook for editing
wb = load_workbook(excel_file_path)
ws = wb.active  # Use the active sheet for writing

# Matrix to store binary vial values for each compound
vial_matrix = []

# Function to extract vial areas and generate binary values for a specific compound
def capture_areas(compound_name, text):
    """
    Capture vial areas for a given compound and convert them to binary (0 or 1).

    Args:
        compound_name (str): The name of the compound.
        text (str): The text containing vial data.

    Returns:
        list: Binary values [0, 0, 0, 0, 0] for the compound's vials.
    """
    # Match the compound's header in the text
    compound_pattern = rf"Compound \d+:  {re.escape(compound_name)}"
    match = re.search(compound_pattern, text)
    
    if not match:
        # If exact match is not found, attempt a fuzzy match
        compound_names_in_text = re.findall(r"Compound \d+:  ([^\n]+)", text)
        closest_match = difflib.get_close_matches(compound_name, compound_names_in_text, n=1, cutoff=0.6)
        
        if closest_match:
            closest_name = closest_match[0]
            compound_pattern = rf"Compound \d+:  {re.escape(closest_name)}"
            match = re.search(compound_pattern, text)
            print(f"Fuzzy match: '{compound_name}' matched with '{closest_name}' in the text.")
        else:
            print(f"'{compound_name}' not found in the text.")
            return [0, 0, 0, 0, 0]
    
    # Initialize binary values for the vials
    areas = {i: 0 for i in range(1, 6)}
    
    if match:
        compound_start = match.end()
        compound_data = text[compound_start:text.find("Compound", compound_start)]
        
        # Match vial numbers and their areas
        vial_pattern = re.compile(r"\s+(\d+)\s+\S+\s+\S+\s+\S+\s+[\d\.]+\s+([\d\.]+)")
        for vial_match in vial_pattern.finditer(compound_data):
            vial_number = int(vial_match.group(1))
            if vial_number <= 5:
                area_value = float(vial_match.group(2))
                areas[vial_number] = 1 if area_value > 6000 else 0
    
    return [areas[i] for i in range(1, 6)]

# Add headers for vials and binary code in the Excel file
header_row = 5  # Assume headers are on row 5
for i in range(5):
    ws.cell(row=header_row, column=compound_column_index + 4 + i, value=f"Vial{i + 1}")

# Add a header for the concatenated binary code
ws.cell(row=header_row, column=compound_column_index + 9, value="Binary Code")

# Process each compound in the Excel file and fill vial binary data
for index, row in df_excel.iterrows():
    compound_name = row[compound_column_index]
    if pd.isna(compound_name) or compound_name == "Compound":
        continue  # Skip empty rows or headers
    
    # Capture vial data for the compound
    vial_values = capture_areas(compound_name, text_data)
    
    # Append binary data to the matrix
    vial_matrix.append(vial_values)
    
    # Write binary vial values to the Excel file
    for i, vial_value in enumerate(vial_values):
        ws.cell(row=index + 1, column=compound_column_index + 4 + i, value=vial_value)

# Save the updated Excel file with binary vial data
wb.save(output_excel_file)
print(f"Updated data saved to '{output_excel_file}'")

# Generate the concatenated binary code and ASCII words
binary_start_column = compound_column_index + 4  # First column for binary vial data
binary_start_row = header_row + 1  # Data starts one row below the headers

with open(output_text_file, "w", encoding="utf-8") as file:
    final_binary_code = ""
    final_word = ""

    # Process each column of vial data to generate binary and ASCII representations
    for col in range(binary_start_column, binary_start_column + 5):
        binary_code = ""
        file.write(f"Vial {col - binary_start_column + 1}:\n")
        
        # Read binary values row by row for the current column
        for row in range(binary_start_row, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value in [0, 1]:  # Ensure only binary values are processed
                binary_code += str(cell_value)
        
        # Convert binary data to ASCII characters
        word = ""
        for i in range(0, len(binary_code), 8):
            binary_segment = binary_code[i:i + 8]
            if len(binary_segment) == 8:
                ascii_char = chr(int(binary_segment, 2))
                word += ascii_char
                final_binary_code += binary_segment
                final_word += ascii_char
            file.write(binary_segment + " ")  # Write binary segments with spaces

        file.write(f"\nASCII Word: {word}\n\n")  # Write the full word for the vial

    # Write the concatenated binary code and ASCII word to the output file
    file.write("\nTotal Binary Code: " + final_binary_code + "\n")
    file.write("Complete ASCII Word: " + final_word + "\n")

print("Total binary code and ASCII word added to the text file.")
